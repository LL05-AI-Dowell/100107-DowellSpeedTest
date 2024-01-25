import csv
from io import BytesIO
import io
import time
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException

import logging

import re
from bs4 import BeautifulSoup
from bs4_web_scraper.scraper import BS4WebScraper
from urllib3.util import parse_url
from urllib.parse import unquote_plus, urljoin, urlsplit, urlunsplit
from difflib import get_close_matches
from concurrent.futures import ThreadPoolExecutor


from .misc import SOCIAL_PLATFORMS, common_address_components


from openpyxl import Workbook, load_workbook
from typing import Optional


class WebsiteInfoScraper:
    """
    Fetches information about a website or webpage.
    """

    engine = BS4WebScraper(log_to_console=False)

    def __init__(self, web_url: str = None, max_search_depth: int = 0, engine: BS4WebScraper = None, browser = None):
        """
        Initializes a WebsiteInfoScraper object.

        :param web_url: The url of the website or webpage to find information about.
        :param engine: The web scraper to use to find information about the website. \
        Should be of type BS4WebScraper.
        :param max_search_depth: The maximum depth to search for information on the website. \
        The higher the depth, the longer it takes to find information. \
        Defaults to 0, which means only the base url is searched.\
        If set to 1, the base url and all the links on the base url are searched. \
        If set to 2, the base url, all the links on the base url and all the links on the links on the base url are searched and so on
        """
        if engine and not isinstance(engine, BS4WebScraper):
            raise TypeError(f"Expected engine to be of type {BS4WebScraper.__name__}, got {type(engine)} instead.")
        self.target = web_url
        self.engine = engine if engine else self.engine
        self.maximum_search_depth = max_search_depth

        
        
        self.browserProfile = webdriver.ChromeOptions()
        
        self.browserProfile.add_argument("--no-sandbox")
        self.browserProfile.add_argument("--headless")  # Run in headless mode
        self.browserProfile.add_argument("--disable-gpu")
        # self.browserProfile.add_experimental_option('prefs', {'intl.accept_languages': 'en,en_US'})
        # self.browserProfile.add_argument("--disable-blink-features=AutomationControlled") 
        # self.browserProfile.add_experimental_option("excludeSwitches", ["enable-automation"]) 
        # self.browserProfile.add_experimental_option("useAutomationExtension", False) 
        

        # if browser.lower() == "chrome":
        # Add similar blocks for other browsers (e.g., Firefox, Edge) if needed
        self.browser = webdriver.Chrome(
            # service=ChromeService(ChromeDriverManager().install()),
            # service=ChromeService(ChromeDriverManager(chrome_type=ChromeType.CHROMIUM).install()), 
            options=self.browserProfile
        ) 
        
        
        # else:
        #     self.browser = webdriver.Firefox

    @staticmethod    
    def normalize_url(url: Optional[str]) -> Optional[str]:
        """
        Normalize the URL by adding 'http://' if the scheme is missing.

        :param url: The URL to normalize.
        :return: Normalized URL.
        """
        if url and not url.startswith(('http://', 'https://')):
            return 'http://' + url
        return url


    def find_website_name(self) -> str | None:
        """
        Find website name.

        :return: Website name if found else None
        """
        base_url = self.engine.get_base_url(self.target)
        set_1 = self.engine.find_all_tags(
            url=base_url,
            target="meta",
            attrs={
                "name": ["application-name", "apple-mobile-web-app-title"],
            },
            depth=self.maximum_search_depth
        )
        set_2 = self.engine.find_all_tags(
            url=base_url,
            target="meta",
            attrs={
                "property": ["og:site_name", "og:title"],
            },
            depth=self.maximum_search_depth
        )
        set_3 = self.engine.find_all_tags(
            url=base_url,
            target="meta",
            attrs={
                "itemprop": ["name"],
            },
            depth=self.maximum_search_depth
        )
        set_4 = self.engine.find_all_tags(
            url=base_url,
            target="title",
            depth=self.maximum_search_depth
        )
        tags = set_1 + set_2 + set_3 + set_4
        results = []
        for tag in tags:
            content = tag.get("content", None) or tag.text.split()[0]
            if content:
                results.append(content)
        names = set(results)
        host = parse_url(base_url).host
        matches = get_close_matches(host, names, n=len(names)//2 or 1, cutoff=0.3)
        return matches[0] if matches else None
    


    def find_emails(self):
        """
        Finds all the emails on the website.

        :return: A list of emails found.
        """
        normalized_url = self.normalize_url(self.target)
        
        try:
            result = self.engine.find_emails(url=normalized_url, depth=self.maximum_search_depth)
            return list(set(result))
        except Exception as e:
            print(f"Error while extracting emails from {normalized_url}: {e}")
            return []


    # def find_emails(self):
    #     """
    #     Finds all the emails on the website.

    #     :return: A list of emails found.
    #     """
    #     result = self.engine.find_emails(url=self.target, depth=self.maximum_search_depth)
    #     return list(set(result))
    

    def find_phone_numbers(self):
        """
        Finds all the phone numbers on the website.

        :return: A list of phone numbers found.
        """
        result = self.engine.find_phone_numbers(url=self.target, depth=self.maximum_search_depth)
        return list(set(result))
        

    def find_links(self):
        """
        Finds all the links on the website.

        :return: A list of urls of the links found.
        """
        links = self.engine.find_links(url=self.target, depth=self.maximum_search_depth)
        links = [link.replace(':///', '://') for link in links]
        return links


    def find_website_logos(self):
        """
        Finds the logos of the website. 
        
        Checks the Favicons, Apple Touch Icons because they are more likely to be the correct ones.
        :return: A list of urls of the logos found.
        """
        logo_tags = self.engine.find_all_tags(
            url=self.target, 
            target="link",
            attrs={
                "rel": ["icon", "shortcut icon", "apple-touch-icon", "apple-touch-icon-precomposed"]
            },
            depth=self.maximum_search_depth
        )
        # Get image urls from meta tags on the websites home page
        image_meta_tags = self.engine.find_all_tags(
            url=self.engine.get_base_url(self.target), 
            target="meta",
            attrs={
                "property": ["og:image", "twitter:image", "twitter:image:src"]
            },
            depth=self.maximum_search_depth
        )
        logo_tags.extend(image_meta_tags)
        logos = [ self.engine.get_tag_rra(tag, download=False) for tag in logo_tags ]
        return list(set(filter(lambda x : bool(x), logos)))
    

    def remove_last_path_segment(self, url):
        parts = list(urlsplit(url))
        path_segments = parts[2].split('/')
        if len(path_segments) > 2:
            path_segments.pop()
        parts[2] = '/'.join(path_segments)
        return urlunsplit(parts)

    def guess_page_url(self, page_name: str):
        """
        Tries to guess the correct URL of a page on the website based on the page name.

        :param page_name: The name of the page to guess the URL of. \
            Can be the name of an external page or a page on the website\
                  as long as it can relate to a URL on the website.
        :return: URL of the page if found, None otherwise.
        """
        base_url = self.engine.get_base_url(self.target)
        links = self.engine.find_links(
            url=base_url,
            depth=self.maximum_search_depth
        )
        links = [link.replace(':///', '://') for link in links]

        # Try and find a close match
        urls = {unquote_plus(link): link for link in links}
        matches = get_close_matches(page_name.lower(), urls.keys(), n=len(urls)//2 or 1, cutoff=0.5)
        match = urls.get(matches[0]) if matches else None
        if match:
            return match

        # If no close match, try and find a link with the page name in it
        for link in links:
            if page_name.lower() in link.lower() or page_name.strip().lower() in link.lower():
                # Remove the last path segment only if there are more than two paths
                modified_link = self.remove_last_path_segment(link)
                return urljoin(base_url, modified_link)

        # If still no match, try and find a link with the page name in its text
        link_tags = self.engine.find_all_tags(
            url=base_url,
            target='a',
            depth=self.maximum_search_depth
        )
        for tag in link_tags:
            if page_name.lower() in tag.text.lower():
                # Remove the last path segment only if there are more than two paths
                modified_link = self.remove_last_path_segment(tag.get("href", ""))
                return urljoin(base_url, modified_link)

        return None

    def guess_pages_urls(self, page_names: list[str]):
        """
        Tries to guess the correct urls of a list of pages on the website based on the page names.

        :param page_names: A list of page names to guess the urls of. \
            Can be the names of external pages or pages on the website\
                  as long as they can related to a url on the website.
        :return: A dictionary of the urls of the pages found.
        """
        r = {}
        def add_result(page_name):
            r[page_name] = self.guess_page_url(page_name)

        with ThreadPoolExecutor() as executor:
            executor.map(add_result, page_names)
        return r
    

    def find_website_social_handle(self, platform_name: str) -> str:
        """
        Finds social media handle for the website on a social media platform.

        This checks the contact and about pages first for social handle,
        because the one on the website details page is more likely to be the correct one.
        If the social handle is not found on those pages, it checks the base url for any link related to the social media platform.

        :param platform_name: The name of the social media platform to find the social handle for.
        :return: The social handle if found, None otherwise.
        """
        # Check contact and about pages first for social handle, 
        # because the one on the website details page is more likely to be the correct one.
        for page_name in ("contact", "about"):
            page_url = self.guess_page_url(page_name)
            if page_url and parse_url(page_url).scheme in ["http", "https"]:
                new_finder = self.__class__(
                    web_url=page_url, 
                    max_search_depth=self.maximum_search_depth,
                    engine = self.engine
                )
                social_url = new_finder.guess_page_url(platform_name)
                if social_url:
                    return social_url
                else:
                    related_links = new_finder.find_links_related_to(platform_name)
                    if related_links:
                        return related_links[0]
            continue
        
        # If url cannot be found in those pages, resort to checking the base url for social handle that matches the website's name
        # This is less likely to be the correct one, but it's better than nothing.
        website_name = self.find_website_name()
        related_links = self.find_links_related_to(platform_name)
        if related_links and website_name:
            matches = get_close_matches(website_name.lower(), related_links, n=len(related_links)//2 or 1, cutoff=0.3)
            if matches:
                return matches[0]
        if related_links:
            return related_links[0]
        return None
    

    def find_website_social_handles(self, platform_names: list[str] = None):
        """
        Finds social media handles for the website on a list of social media platforms.

        :param platform_names: A list of social media platforms to find the social handles for. If None,\
              all the supported social media platforms are searched for.
        :return: A dictionary of the social media handles found.
        """
        website_base_url = self.engine.get_base_url(self.target)
        platform_names = list(SOCIAL_PLATFORMS.keys()) if not platform_names else [ name.lower() for name in platform_names ]
        if not platform_names:
            # Remove the website from the list of social platform url to search for. 
            # If the website its self is a social platform,
            for index, (_, value) in enumerate(SOCIAL_PLATFORMS.items()):
                if website_base_url in value:
                    platform_names.pop(index)
                    break
        r = {}
        def add_result(platform_name):
            r[platform_name] = self.find_website_social_handle(platform_name)

        with ThreadPoolExecutor() as executor:
            executor.map(add_result, platform_names)
        return r or None
    

    def find_links_related_to(self, _s: str | list[str], links: list[str] = None):
        """
        Similar to `guess_pages_urls` method, but is more general and can return single or multiple results.
        For more accurate multiple results, use `guess_pages_urls` method instead.
        For more accurate single results, use `guess_page_url` method instead.
        For better performance, use this method.

        :param _s: The string or list of strings to search for in the links.
        :param links: A list of links to search in. If None, the links on the website are searched.
        :return: A list of urls of the links found or a dictionary of each result if `_s` is a list.
        """
        links = self.find_links() if not links else links

        def get_matches(string: str):
            return [ link for link in links if string.lower() in link.lower() ] or None
        
        if isinstance(_s, str):
            return get_matches(_s)
        return { _s_ : get_matches(_s_) for _s_ in _s }


    def find_all_social_media_links(self, platform_names: list[str] = None):
        """
        Finds all the social media links on the website.

        :param platform_names: A list of social media platforms to find the social media links for. If None,\
              all the supported social media platforms are searched for.
        :return: A dictionary of the social media links found.
        """
        website_base_url = self.engine.get_base_url(self.target)
        platform_names = list(SOCIAL_PLATFORMS.keys()) if not platform_names else [ name.lower() for name in platform_names ]
        if not platform_names:
            # Remove the website from the list of social platform url to search for. 
            # If the website its self is a social platform,
            for index, (_, value) in enumerate(SOCIAL_PLATFORMS.items()):
                if website_base_url in value:
                    platform_names.pop(index)
                    break
        return self.find_links_related_to(platform_names)
    
    
    def find_addresses(self):
        """Finds all addresses on website. Results are not guaranteed to be accurate."""
        address_regex = r"(\d{2,5}[-;\.\|:]*)+(\s*[\w\.\s,]+){1,3}[\.|\s]*(\w{2,3})?[\.|\s]*(\d{5}|\w{2,3})?"
        results = self.engine.find_pattern(
            url=self.target,
            regex=address_regex,
        )
        addresses = []
        for result in results:
            for c in common_address_components:
                if c in result:
                    result = re.sub(r"\n+", " ", result).strip()
                    addresses.append(result)
                    break
        return addresses or None
    
    def get_page(self, web_url):
        self.browser.get(web_url)
        # Wait for the page to load (you can adjust the timeout as needed)
        WebDriverWait(self.browser, 10).until(EC.presence_of_element_located((By.TAG_NAME, 'form')))

        # Parse the HTML content of the page
        soup = BeautifulSoup(self.browser.page_source, 'html.parser')
        return soup
    
    def extract_form_data(self, form):
        form_data = {}
        form_fields = form.find_all(['input', 'textarea'])

        for field in form_fields:
            field_name = field.get('name')
            field_type = field.get('type')
            field_id = field.get('id')

            is_hidden = field.get('type') == 'hidden'
            is_submit = field.get('type') == 'submit'

            # Exclude fields named "submit" and hidden fields
            if is_submit or is_hidden:
                continue
                
        
            if field_type and field_name:
                form_data[field_name] = field_type
            
            elif field.name == 'textarea' and field_name:
                form_data[field_name] = 'textarea'
            elif field.name == 'textarea' and field_id:
                form_data[field_id] = 'textarea'
            elif field_id and not field_name:
                form_data[field_id] = field_type
            else:
                form_data[field_type] = field_type

        return form_data

 
    def scrape_contact_us_page(self, web_url):
        try:
            soup = self.get_page(web_url)

            # for debugging purposes
            html_str = soup.prettify()
            with open("soup.html", "w", encoding="utf-8") as file:
                file.write(html_str)
                
            form_elements = soup.find_all('form')

            # If there's only one form, return a single dictionary
            if len(form_elements) == 1:
                form_data = self.extract_form_data(form_elements[0])
                if form_data:
                    return form_data
                else:
                    raise Exception("No Form Fields found on the Contact Us Form.")
                
            # If there are multiple forms, return a list of forms
            elif len(form_elements) > 1:

                form_data_list = [self.extract_form_data(form) for form in form_elements]

                # Append an index to each form data dictionary
                # form_data_list_with_index = [{"form_index": index, **data} for index, data in enumerate(form_data_list)]

                # Remove duplicates while preserving the index
                unique_form_data_list = []

                for data in form_data_list:
                    if data not in unique_form_data_list:
                        unique_form_data_list.append(data)
                    
                if unique_form_data_list:
                    return unique_form_data_list
                else:
                    raise Exception("No Form Fields found on the Contact Us Forms.")
            else:
                raise Exception("Form(s) not found on the Contact Us page")
        except Exception as e:
            raise Exception(f"An error occurred: {str(e)}")
        
        finally:
            self.browser.quit()
            

    def save_form_data_to_excel(self, form_data, file_type):
        if not form_data:
            exception = f"No form's found"
            logging.warning(exception)
            raise Exception(exception)

        if isinstance(form_data, dict):
            form_data = [form_data]

        if file_type == "xlsx":
            wb = Workbook()

            for data in form_data:
                sheet = wb.create_sheet(title=f"Form")  # Create a new sheet for each form
                header = list(data.keys())
                sheet.append(header)  # Add headers

                # Add the form data as a row
                data_row = [data[key] for key in data.keys()]
                sheet.append(data_row)

            del wb["Sheet"]  # Remove the default sheet

            output = io.BytesIO()
            wb.save(output)

            # Reset the stream position to the beginning
            output.seek(0)

            return output.getvalue()

        elif file_type == "csv":
            output = io.StringIO()

            for index, data in enumerate(form_data, start=1):
                csv_writer = csv.writer(output)
                
                # Write the CSV header
                header = list(data.keys())
                csv_writer.writerow(header)
                
                # Write the CSV data
                csv_writer.writerow([data[key] for key in data.keys()])
                
                output.write("\n")  # Add a newline to separate sheets

            return output.getvalue()

        else:
            raise Exception("Unsupported file type. Please choose 'csv' or 'xlsx'.")

    def extract_excel_data(self, file):
        try:
            workbook = load_workbook(file, read_only=True)
            form_data = []

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # Assuming the first row contains headers
                headers = [cell for cell in next(sheet.iter_rows(values_only=True))]

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    data = dict(zip(headers, row))
                    form_data.append(data)

            logging.info(form_data)
            return form_data

        except Exception as e:
            raise Exception(f"Error extracting data from XLSX file: {e}")


    def submit_contact_form_selenium(self, contact_us_links, form_data_list):
        try:
            # Open the contact_us_link
            response_data = []
            successful_submissions = 0

            for contact_us_link in contact_us_links:
                self.browser.get(contact_us_link)
                WebDriverWait(self.browser, 10).until(EC.presence_of_element_located((By.TAG_NAME, 'form')))
                form_elements = self.browser.find_elements(By.TAG_NAME, 'form')

                for form_data in form_data_list:
                   
                    for form in form_elements:
                        for field_name, value in form_data.items():
                            try:
                                input_field = form.find_element(By.NAME, field_name)
                            except NoSuchElementException:
                                try:
                                    input_field = form.find_element(By.ID, field_name)
                                except NoSuchElementException:
                                    continue  # Skip to the next iteration if the field is not found
                            
                            try:
                                input_field.send_keys(value)
                            except:
                                continue
                        try:
                            time.sleep(4)
                            form.submit()
                            successful_submissions += 1
                            response = f"Form {successful_submissions} submitted successfully."
                            response_data.append(response)
                        except Exception as e:
                            response = f"Error submitting form {str(e)}"
                            response_data.append({"error": response})
                            continue
            time.sleep(4)
            # Close the WebDriver outside the loop
            self.browser.quit()

            return response_data
        except Exception as e:
            raise Exception(e)
            
    